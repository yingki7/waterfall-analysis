import streamlit as st
import pandas as pd
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import io

st.set_page_config(page_title="广告瀑布流分析工具", layout="wide")

st.title("📊 广告单元瀑布流分析工具")
st.markdown("上传CSV文件，自动生成Top 5国家的广告单元分析报告")


def apply_excel_styling(worksheet, data_df, start_row):
    """应用Excel样式（包含条件格式 + 优化列宽）"""
    
    content_font = Font(name='等线', size=11)
    header_font = Font(name='等线', bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    even_fill = PatternFill(start_color="F5F9FF", end_color="F5F9FF", fill_type="solid")
    odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    columns = data_df.columns.tolist()
    
    # 找出需要条件格式的列
    earnings_pct_idx = None
    impressions_pct_idx = None
    for idx, col_name in enumerate(columns, 1):
        if col_name == 'Earnings %':
            earnings_pct_idx = idx
        elif col_name == 'Impressions %':
            impressions_pct_idx = idx
    
    # Excel条件格式风格的颜色
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # 设置表头
    worksheet.row_dimensions[start_row].height = 30
    for col_idx, col_name in enumerate(columns, 1):
        cell = worksheet.cell(row=start_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    # 设置数据行
    for row_idx in range(start_row + 1, start_row + len(data_df) + 1):
        worksheet.row_dimensions[row_idx].height = 22
        
        base_fill = even_fill if (row_idx - start_row - 1) % 2 == 1 else odd_fill
        
        for col_idx, col_name in enumerate(columns, 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.font = content_font
            cell.fill = base_fill
            cell.border = thin_border
            
            if col_name == 'Rank':
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name == 'Ad Unit':
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            
            if col_name == 'Earnings (USD)':
                cell.number_format = '"$"#,##0.00'
            elif col_name == 'eCPM (USD)':
                cell.number_format = '"$"#,##0.00'
            elif col_name in ['Earnings %', 'Match Rate (%)', 'Impressions %', 'Requests %']:
                cell.number_format = '0.00"%"'
            elif col_name in ['Impressions', 'Requests']:
                cell.number_format = '#,##0'
        
        # 条件格式：Earnings %
        if earnings_pct_idx:
            cell = worksheet.cell(row=row_idx, column=earnings_pct_idx)
            value = cell.value
            if value is not None:
                if value > 10:
                    cell.fill = green_fill
                elif value < 2:
                    cell.fill = red_fill
        
        # 条件格式：Impressions %
        if impressions_pct_idx:
            cell = worksheet.cell(row=row_idx, column=impressions_pct_idx)
            value = cell.value
            if value is not None:
                if value > 10:
                    cell.fill = green_fill
                elif value < 2:
                    cell.fill = red_fill
    
    # 优化列宽
    column_widths = {
        'Rank': 6,
        'Ad Unit': 30,
        'Earnings (USD)': 16,
        'Earnings %': 14,
        'Impressions': 14,
        'Impressions %': 14,
        'Requests': 14,
        'Requests %': 14,
        'eCPM (USD)': 14,
        'Match Rate (%)': 16
    }
    
    for col_idx, col_name in enumerate(columns, 1):
        column_letter = get_column_letter(col_idx)
        
        if col_name in column_widths:
            base_width = column_widths[col_name]
        else:
            base_width = max(len(col_name) + 3, 12)
        
        max_data_length = 0
        for row in range(start_row + 1, min(start_row + len(data_df) + 1, start_row + 200)):
            cell_value = worksheet.cell(row=row, column=col_idx).value
            if cell_value is not None:
                if col_name in ['Earnings (USD)', 'eCPM (USD)']:
                    length = len(f"${cell_value:,.2f}") if isinstance(cell_value, (int, float)) else len(str(cell_value))
                elif col_name in ['Earnings %', 'Impressions %', 'Requests %', 'Match Rate (%)']:
                    length = len(f"{cell_value:.2f}%") if isinstance(cell_value, (int, float)) else len(str(cell_value))
                elif col_name in ['Impressions', 'Requests']:
                    length = len(f"{cell_value:,.0f}") if isinstance(cell_value, (int, float)) else len(str(cell_value))
                else:
                    length = len(str(cell_value))
                
                if length > max_data_length:
                    max_data_length = min(length, 50)
        
        header_length = len(col_name) + 2
        final_width = max(header_length, max_data_length + 2, base_width)
        
        if col_name == 'Ad Unit':
            final_width = min(final_width, 30)
        else:
            final_width = min(final_width, 22)
        
        worksheet.column_dimensions[column_letter].width = final_width


def process_csv(file_content):
    df = pd.read_csv(io.BytesIO(file_content), sep='\t', encoding='utf-16')
    
    numeric_columns = ['Estimated earnings (USD)', 'Observed eCPM (USD)', 'Requests',
                      'Matched requests', 'Show rate', 'Impressions', 'CTR', 'Clicks',
                      'Ads ARPV (USD)', 'Ads ARPU (USD)', 'Ad viewers (AV)', 'Active users (AU)',
                      'Ad viewer rate', 'Imps / AV', 'Imps / AU', 'Ads ARPDAV (USD)', 'Ads ARPDAU (USD)',
                      'DAV', 'DAU', 'Daily ad viewer rate', 'IMPDAV', 'IMPDAU', 'Ad load latency',
                      'Bid requests', 'Bids in auction (%)', 'Bids in auction', 'Win rate', 'Winning bids']
    for col in numeric_columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    
    df['Match rate'] = pd.to_numeric(df['Match rate'].astype(str).str.replace('%', '', regex=False), errors='coerce')
    
    country_earnings = df.groupby('Country')['Estimated earnings (USD)'].sum().sort_values(ascending=False)
    top_5_countries = country_earnings.head(5).index.tolist()
    
    interstitial_native_df = df[
        df['Ad unit'].str.contains('Interstitial|Native|Full|Inter', case=False, na=False) &
        ~df['Ad unit'].str.contains('banner', case=False, na=False)
    ].copy()
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        summary_data = []
        for country in top_5_countries:
            country_data = interstitial_native_df[interstitial_native_df['Country'] == country].copy()
            if len(country_data) > 0:
                total_earnings = country_data['Estimated earnings (USD)'].sum()
                total_impressions = country_data['Impressions'].sum()
                total_requests = country_data['Requests'].sum()
                
                result_df = pd.DataFrame({
                    'Ad Unit': country_data['Ad unit'],
                    'Earnings (USD)': country_data['Estimated earnings (USD)'].round(2),
                    'Earnings %': (country_data['Estimated earnings (USD)'] / total_earnings * 100).round(2),
                    'Impressions': country_data['Impressions'].fillna(0).astype(int),
                    'Impressions %': (country_data['Impressions'] / total_impressions * 100).round(2),
                    'Requests': country_data['Requests'].fillna(0).astype(int),
                    'Requests %': (country_data['Requests'] / total_requests * 100).round(2),
                    'eCPM (USD)': country_data['Observed eCPM (USD)'].round(2),
                    'Match Rate (%)': country_data['Match rate'].round(2)
                }).fillna(0).sort_values('eCPM (USD)', ascending=False).reset_index(drop=True)
                result_df.insert(0, 'Rank', range(1, len(result_df) + 1))
                
                app_name = country_data['App'].mode()[0] if 'App' in country_data.columns and not country_data['App'].mode().empty else "Unknown"
                sheet_name = re.sub(r'[\\/*?:"<>|]', '', country)[:31]
                start_row = 6
                result_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row - 1, index=False)
                worksheet = writer.sheets[sheet_name]
                
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.font = Font(name='等线', size=10)
                
                app_cell = worksheet.cell(row=1, column=1, value=f" App: {app_name}")
                app_cell.font = Font(name='等线', bold=True, size=14)
                app_cell.alignment = Alignment(horizontal="left", vertical="center")
                worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
                worksheet.row_dimensions[1].height = 28
                
                title_cell = worksheet.cell(row=2, column=1, value=f"📊 {country} - 广告单元分析报告")
                title_cell.font = Font(name='等线', bold=True, size=16)
                title_cell.alignment = Alignment(horizontal="left", vertical="center")
                worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)
                worksheet.row_dimensions[2].height = 30
                
                stats = [
                    (f"总收益: ${total_earnings:,.2f}", 3, 1),
                    (f"总展示次数: {total_impressions:,.0f}", 3, 3),
                    (f"总请求数: {total_requests:,.0f}", 3, 5),
                    (f"Ad Units数量: {len(country_data)}", 3, 7),
                    (f"平均eCPM: ${result_df['eCPM (USD)'].mean():.2f}", 4, 1),
                    (f"平均Match Rate: {result_df['Match Rate (%)'].mean():.2f}%", 4, 3),
                    (f"最高eCPM: ${result_df['eCPM (USD)'].max():.2f}", 4, 5),
                    (f"最低eCPM: ${result_df['eCPM (USD)'].min():.2f}", 4, 7),
                ]
                for text, row, col in stats:
                    cell = worksheet.cell(row=row, column=col, value=text)
                    cell.font = Font(name='等线', size=11, bold=True)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                
                worksheet.row_dimensions[3].height = 22
                worksheet.row_dimensions[4].height = 22
                
                for col in range(1, 10):
                    cell = worksheet.cell(row=5, column=col, value="")
                    cell.border = Border(bottom=Side(style='medium', color='4472C4'))
                
                apply_excel_styling(worksheet, result_df, start_row)
                worksheet.freeze_panes = f'A{start_row + 1}'
                
                summary_data.append({
                    'App Name': app_name,
                    'Country': country,
                    'Total Earnings (USD)': round(total_earnings, 2),
                    'Total Impressions': int(total_impressions),
                    'Total Requests': int(total_requests),
                    'Number of Ad Units': len(country_data),
                    'Avg eCPM (USD)': round(result_df['eCPM (USD)'].mean(), 2),
                    'Avg Match Rate (%)': round(result_df['Match Rate (%)'].mean(), 2),
                    'Max eCPM (USD)': round(result_df['eCPM (USD)'].max(), 2),
                    'Min eCPM (USD)': round(result_df['eCPM (USD)'].min(), 2)
                })
        
        if summary_data:
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False, startrow=2)
            worksheet = writer.sheets['Summary']
            
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.font = Font(name='等线', size=10)
            
            worksheet.cell(row=1, column=1, value="📈 Top 5国家广告单元汇总分析").font = Font(name='等线', bold=True, size=16)
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
            worksheet.row_dimensions[1].height = 30
            
            header_row = 3
            worksheet.row_dimensions[header_row].height = 25
            header_font = Font(name='等线', bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            for col_idx, col_name in enumerate(summary_df.columns, 1):
                cell = worksheet.cell(row=header_row, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            content_font = Font(name='等线', size=11)
            even_fill = PatternFill(start_color="F5F9FF", end_color="F5F9FF", fill_type="solid")
            thin_border = Border(left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
                                top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0'))
            for row_idx in range(header_row + 1, len(summary_df) + header_row + 1):
                worksheet.row_dimensions[row_idx].height = 20
                row_fill = even_fill if (row_idx - header_row - 1) % 2 == 1 else PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                for col_idx, col_name in enumerate(summary_df.columns, 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.font = content_font
                    cell.fill = row_fill
                    cell.border = thin_border
                    if col_name in ['Total Earnings (USD)', 'Avg eCPM (USD)', 'Max eCPM (USD)', 'Min eCPM (USD)']:
                        cell.number_format = '"$"#,##0.00'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif col_name == 'Avg Match Rate (%)':
                        cell.number_format = '0.00"%"'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif col_name in ['Total Impressions', 'Total Requests']:
                        cell.number_format = '#,##0'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
            
            worksheet.freeze_panes = f'A{header_row + 1}'
            
            summary_column_widths = {
                'App Name': 20,
                'Country': 18,
                'Total Earnings (USD)': 20,
                'Total Impressions': 18,
                'Total Requests': 16,
                'Number of Ad Units': 16,
                'Avg eCPM (USD)': 16,
                'Avg Match Rate (%)': 18,
                'Max eCPM (USD)': 16,
                'Min eCPM (USD)': 16
            }
            for col_idx, col_name in enumerate(summary_df.columns, 1):
                column_letter = get_column_letter(col_idx)
                base_width = summary_column_widths.get(col_name, 15)
                max_length = len(col_name) + 2
                for row in range(header_row + 1, len(summary_df) + header_row + 1):
                    cell_value = worksheet.cell(row=row, column=col_idx).value
                    if cell_value is not None:
                        length = len(str(cell_value))
                        if length > max_length:
                            max_length = min(length, 25)
                worksheet.column_dimensions[column_letter].width = max(max_length + 2, base_width)
    
    output.seek(0)
    return output


# ========== Streamlit 界面 ==========
uploaded_file = st.file_uploader("📁 上传CSV文件", type=['csv', 'txt'])

if uploaded_file is not None:
    with st.spinner('🔄 正在分析数据，请稍候...'):
        try:
            df_preview = pd.read_csv(io.BytesIO(uploaded_file.read()), sep='\t', encoding='utf-16')
            with st.expander("📋 数据预览（前10行）"):
                st.dataframe(df_preview.head(10))
            
            uploaded_file.seek(0)
            excel_data = process_csv(uploaded_file.read())
            st.success("✅ 分析完成！")
            
            app_name = "ad_analysis"
            try:
                uploaded_file.seek(0)
                df = pd.read_csv(io.BytesIO(uploaded_file.read()), sep='\t', encoding='utf-16')
                if 'App' in df.columns and not df['App'].mode().empty:
                    app_name = re.sub(r'[\\/*?:"<>|]', '', str(df['App'].mode()[0]))[:50]
            except:
                pass
            
            st.download_button(
                label="📥 下载Excel报告",
                data=excel_data,
                file_name=f"{app_name}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            st.info("💡 报告已生成，点击上方按钮下载")
            st.markdown("""
            **📌 颜色说明：**
            - 🟢 绿色：收益占比/展示占比 > 10%
            - 🔴 红色：收益占比/展示占比 < 2%
            - ⚪ 白色/浅蓝：正常范围 (2% ~ 10%)
            """)
            
        except Exception as e:
            st.error(f"❌ 处理失败: {str(e)}")
            st.exception(e)

st.markdown("---")
st.caption("💡 数据格式要求：UTF-16编码的Tab分隔CSV文件，必须包含 App, Country, Ad unit 等列")
